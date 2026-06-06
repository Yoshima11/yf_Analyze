"""
================================================================================
DCF ANALYZER PROFESIONAL - VERSIÓN CORREGIDA
================================================================================
"""

import concurrent.futures
import time
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
import yfinance as yf

try:
    from openpyxl.styles import PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

warnings.filterwarnings("ignore")

# ============================================
# CONFIGURACIÓN DE FILTROS
# ============================================

FILTER_CONFIG = {
    "min_margin": -30,
    "max_margin": 100,
    "show_only_buy": False,
    "min_cagr": -10,
    "max_debt_to_ebitda": 8,
    "min_quality_score": 3,
    "max_workers": 5,
    "save_checkpoint_every": 50,
}

# ============================================
# SECTORES QUE SON BANCOS
# ============================================

BANK_SECTORS = [
    "Bancos/Custodia",
    "Bancos",
    "Banca Inversión",
    "Bancos Japoneses",
    "Bancos UK",
    "Bancos Holanda",
    "Bancos España",
    "Bancos Brasil",
    "Bancos Corea",
    "Bancos Canadá",
    "Bancos Argentina",
    "Bancos India",
    "Bancos US",
    "Bancos Japón",
    "Broker",
    "Bancos Australia",
]

# ============================================
# EXCLUIR ETFs
# ============================================

ETF_TICKERS = ["SPY", "QQQ", "IWM", "DIA", "XLF", "XLE"]

# ============================================
# TU DICCIONARIO DE ACCIONES
# ============================================

ACCIONES_US_TOP = {
    "AAPL": {"sector": "Tecnología", "prioridad": "muy_alta"},
    "MSFT": {"sector": "Cloud/AI", "prioridad": "muy_alta"},
    "NVDA": {"sector": "Semiconductores/AI", "prioridad": "muy_alta"},
    "GOOGL": {"sector": "Internet/Publicidad", "prioridad": "alta"},
    "META": {"sector": "Social Media", "prioridad": "alta"},
    "AMZN": {"sector": "E-commerce/Cloud", "prioridad": "alta"},
    "TSLA": {"sector": "Automoción/Tech", "prioridad": "muy_alta"},
    "ADBE": {"sector": "Software", "prioridad": "alta"},
    "AMD": {"sector": "Semiconductores", "prioridad": "alta"},
    "AVGO": {"sector": "Semiconductores", "prioridad": "alta"},
    "CRM": {"sector": "Software Cloud", "prioridad": "media"},
    "INTC": {"sector": "Semiconductores", "prioridad": "media"},
    "MU": {"sector": "Semiconductores", "prioridad": "media"},
    "QCOM": {"sector": "Semiconductores", "prioridad": "media"},
    "TXN": {"sector": "Semiconductores", "prioridad": "media"},
    "LRCX": {"sector": "Equipo Semiconductores", "prioridad": "media"},
    "AMAT": {"sector": "Equipo Semiconductores", "prioridad": "media"},
    "ORCL": {"sector": "Software/Database", "prioridad": "alta"},
    "IBM": {"sector": "Tecnología/Consultoría", "prioridad": "media"},
    "HPQ": {"sector": "Hardware", "prioridad": "baja"},
    "GLW": {"sector": "Tecnología Industrial", "prioridad": "baja"},
    "VRSN": {"sector": "Internet", "prioridad": "baja"},
    "PANW": {"sector": "Ciberseguridad", "prioridad": "alta"},
    "SNOW": {"sector": "Cloud/Datos", "prioridad": "media"},
    "PLTR": {"sector": "Analítica de Datos", "prioridad": "media"},
    "SQ": {"sector": "Fintech", "prioridad": "media"},
    "PYPL": {"sector": "Fintech", "prioridad": "alta"},
    "UBER": {"sector": "Transporte/Tech", "prioridad": "media"},
    "SHOP": {"sector": "E-commerce", "prioridad": "media"},
    "EBAY": {"sector": "E-commerce", "prioridad": "media"},
    "ETSY": {"sector": "E-commerce", "prioridad": "baja"},
    "SNAP": {"sector": "Social Media", "prioridad": "baja"},
    "TWTR": {"sector": "Social Media", "prioridad": "baja"},
    "SPOT": {"sector": "Streaming", "prioridad": "media"},
    "NFLX": {"sector": "Streaming", "prioridad": "alta"},
    "DIS": {"sector": "Entretenimiento", "prioridad": "alta"},
    "ROKU": {"sector": "Streaming", "prioridad": "baja"},
    "ZM": {"sector": "Comunicaciones", "prioridad": "baja"},
    "DOCU": {"sector": "Software", "prioridad": "baja"},
    "MSTR": {"sector": "Software/Bitcoin", "prioridad": "media"},
    "CSCO": {"sector": "Tecnología/Redes", "prioridad": "media"},
    "ADI": {"sector": "Semiconductores", "prioridad": "media"},
    "ADP": {"sector": "Servicios Empresariales", "prioridad": "alta"},
    "GRMN": {"sector": "Tecnología GPS", "prioridad": "media"},
    "MSI": {"sector": "Telecom Equipos", "prioridad": "media"},
    "RBLX": {"sector": "Gaming/Metaverso", "prioridad": "media"},
    "SNA": {"sector": "Herramientas", "prioridad": "baja"},
    "XRX": {"sector": "Tecnología Oficina", "prioridad": "baja"},
    "YELP": {"sector": "Internet/Reviews", "prioridad": "baja"},
    "SATL": {"sector": "Satélites/Tech", "prioridad": "media"},
    "BK": {"sector": "Bancos/Custodia", "prioridad": "muy_alta"},
    "GS": {"sector": "Banca Inversión", "prioridad": "muy_alta"},
    "JPM": {"sector": "Bancos", "prioridad": "alta"},
    "BAC": {"sector": "Bancos", "prioridad": "alta"},
    "C": {"sector": "Bancos", "prioridad": "media"},
    "WFC": {"sector": "Bancos", "prioridad": "media"},
    "USB": {"sector": "Bancos", "prioridad": "media"},
    "MS": {"sector": "Banca Inversión", "prioridad": "alta"},
    "AXP": {"sector": "Tarjetas de Crédito", "prioridad": "alta"},
    "V": {"sector": "Medios de Pago", "prioridad": "muy_alta"},
    "MA": {"sector": "Medios de Pago", "prioridad": "muy_alta"},
    "COIN": {"sector": "Cripto/Fintech", "prioridad": "media"},
    "SPGI": {"sector": "Ratings/Finanzas", "prioridad": "alta"},
    "MMC": {"sector": "Seguros/Corretaje", "prioridad": "media"},
    "AIG": {"sector": "Seguros", "prioridad": "media"},
    "MET": {"sector": "Seguros", "prioridad": "baja"},
    "PRU": {"sector": "Seguros", "prioridad": "baja"},
    "SCHW": {"sector": "Broker", "prioridad": "alta"},
    "BLK": {"sector": "Gestión Activos", "prioridad": "alta"},
    "BRK/B": {"sector": "Conglomerado", "prioridad": "muy_alta"},
    "TRV": {"sector": "Seguros US", "prioridad": "media"},
    "UPST": {"sector": "Fintech Préstamos", "prioridad": "media"},
    "MUFG": {"sector": "Bancos Japoneses", "prioridad": "alta"},
    "MFG": {"sector": "Bancos Japoneses", "prioridad": "alta"},
    "NMR": {"sector": "Bancos Japoneses", "prioridad": "alta"},
    "SMFG": {"sector": "Bancos Japoneses", "prioridad": "media"},
    "HSBC": {"sector": "Bancos UK", "prioridad": "alta"},
    "BCS": {"sector": "Bancos UK", "prioridad": "media"},
    "LYG": {"sector": "Bancos UK", "prioridad": "media"},
    "ING": {"sector": "Bancos Holanda", "prioridad": "media"},
    "BBVA": {"sector": "Bancos España", "prioridad": "alta"},
    "SAN": {"sector": "Bancos España", "prioridad": "media"},
    "ITUB": {"sector": "Bancos Brasil", "prioridad": "media"},
    "BBD": {"sector": "Bancos Brasil", "prioridad": "media"},
    "BSBR": {"sector": "Bancos Brasil", "prioridad": "baja"},
    "KB": {"sector": "Bancos Corea", "prioridad": "media"},
    "SHG": {"sector": "Bancos Corea", "prioridad": "media"},
    "RY": {"sector": "Bancos Canadá", "prioridad": "media"},
    "TD": {"sector": "Bancos Canadá", "prioridad": "media"},
    "WBK": {"sector": "Bancos Australia", "prioridad": "baja"},
    "GGAL": {"sector": "Bancos Argentina", "prioridad": "alta"},
    "BBAR": {"sector": "Bancos Argentina", "prioridad": "alta"},
    "BMA": {"sector": "Bancos Argentina", "prioridad": "alta"},
    "SUPV": {"sector": "Bancos Argentina", "prioridad": "media"},
    "BYMA": {"sector": "Finanzas Argentina", "prioridad": "media"},
    "VIST": {"sector": "Energía (Vaca Muerta)", "prioridad": "muy_alta"},
    "YPF": {"sector": "Energía Argentina", "prioridad": "alta"},
    "PAM": {"sector": "Energía Argentina", "prioridad": "alta"},
    "CEPU": {"sector": "Energía Argentina", "prioridad": "media"},
    "XOM": {"sector": "Petróleo & Gas", "prioridad": "alta"},
    "CVX": {"sector": "Petróleo & Gas", "prioridad": "alta"},
    "BP": {"sector": "Petróleo & Gas", "prioridad": "media"},
    "SHEL": {"sector": "Petróleo & Gas", "prioridad": "alta"},
    "TOT": {"sector": "Petróleo & Gas", "prioridad": "media"},
    "PSX": {"sector": "Refinación", "prioridad": "alta"},
    "OXY": {"sector": "Petróleo & Gas", "prioridad": "media"},
    "COP": {"sector": "Petróleo & Gas", "prioridad": "media"},
    "SLB": {"sector": "Servicios Petróleo", "prioridad": "media"},
    "HAL": {"sector": "Servicios Petróleo", "prioridad": "media"},
    "FSLR": {"sector": "Energía Solar", "prioridad": "alta"},
    "NEE": {"sector": "Energía Eléctrica", "prioridad": "alta"},
    "ENPH": {"sector": "Energía Solar", "prioridad": "media"},
    "PBR": {"sector": "Petróleo Brasil", "prioridad": "alta"},
    "TS": {"sector": "Energía Argentina", "prioridad": "alta"},
    "TGS": {"sector": "Gas Argentina", "prioridad": "media"},
    "TGNO4": {"sector": "Gas Argentina", "prioridad": "media"},
    "TRAN": {"sector": "Energía Argentina", "prioridad": "media"},
    "EDN": {"sector": "Energía Argentina", "prioridad": "media"},
    "GPRK": {"sector": "Petróleo Colombia", "prioridad": "media"},
    "BHP": {"sector": "Minería", "prioridad": "alta"},
    "RIO": {"sector": "Minería", "prioridad": "alta"},
    "VALE": {"sector": "Minería/Hierro", "prioridad": "alta"},
    "NEM": {"sector": "Minería/Oro", "prioridad": "media"},
    "GOLD": {"sector": "Minería/Oro", "prioridad": "media"},
    "NUE": {"sector": "Acero", "prioridad": "alta"},
    "BA": {"sector": "Aeroespacial", "prioridad": "alta"},
    "RTX": {"sector": "Aeroespacial/Defensa", "prioridad": "alta"},
    "LMT": {"sector": "Aeroespacial/Defensa", "prioridad": "alta"},
    "GE": {"sector": "Industrial", "prioridad": "alta"},
    "HON": {"sector": "Industrial", "prioridad": "alta"},
    "MMM": {"sector": "Industrial", "prioridad": "alta"},
    "CAT": {"sector": "Maquinaria Pesada", "prioridad": "alta"},
    "DE": {"sector": "Maquinaria Agrícola", "prioridad": "media"},
    "PCAR": {"sector": "Camiones", "prioridad": "media"},
    "FDX": {"sector": "Logística", "prioridad": "alta"},
    "UPS": {"sector": "Logística", "prioridad": "alta"},
    "UNP": {"sector": "Ferrocarriles", "prioridad": "media"},
    "CSX": {"sector": "Ferrocarriles", "prioridad": "media"},
    "NSC": {"sector": "Ferrocarriles", "prioridad": "media"},
    "JNJ": {"sector": "Salud/Consumo", "prioridad": "alta"},
    "PFE": {"sector": "Farmacéutica", "prioridad": "alta"},
    "MRK": {"sector": "Farmacéutica", "prioridad": "alta"},
    "ABBV": {"sector": "Farmacéutica", "prioridad": "alta"},
    "BMY": {"sector": "Farmacéutica", "prioridad": "media"},
    "LLY": {"sector": "Farmacéutica", "prioridad": "muy_alta"},
    "AMGN": {"sector": "Biotecnología", "prioridad": "alta"},
    "GILD": {"sector": "Biotecnología", "prioridad": "media"},
    "BIIB": {"sector": "Biotecnología", "prioridad": "media"},
    "REGN": {"sector": "Biotecnología", "prioridad": "media"},
    "MDT": {"sector": "Dispositivos Médicos", "prioridad": "alta"},
    "ABT": {"sector": "Dispositivos Médicos", "prioridad": "alta"},
    "TMO": {"sector": "Equipo Científico", "prioridad": "alta"},
    "DHR": {"sector": "Equipo Científico", "prioridad": "alta"},
    "UNH": {"sector": "Seguros de Salud", "prioridad": "alta"},
    "CVS": {"sector": "Farmacias/Salud", "prioridad": "media"},
    "CI": {"sector": "Seguros de Salud", "prioridad": "media"},
    "NVS": {"sector": "Farmacéutica Suiza", "prioridad": "alta"},
    "AZN": {"sector": "Farmacéutica UK", "prioridad": "alta"},
    "GSK": {"sector": "Farmacéutica UK", "prioridad": "media"},
    "KO": {"sector": "Bebidas", "prioridad": "muy_alta"},
    "PEP": {"sector": "Bebidas/Snacks", "prioridad": "alta"},
    "KMB": {"sector": "Productos Consumo", "prioridad": "alta"},
    "PG": {"sector": "Productos Consumo", "prioridad": "alta"},
    "CL": {"sector": "Productos Consumo", "prioridad": "media"},
    "MO": {"sector": "Tabaco", "prioridad": "media"},
    "WMT": {"sector": "Retail/Masivo", "prioridad": "alta"},
    "COST": {"sector": "Retail/Masivo", "prioridad": "alta"},
    "TGT": {"sector": "Retail", "prioridad": "media"},
    "HD": {"sector": "Mejoramiento Hogar", "prioridad": "alta"},
    "LOW": {"sector": "Mejoramiento Hogar", "prioridad": "media"},
    "MCD": {"sector": "Restaurantes", "prioridad": "alta"},
    "SBUX": {"sector": "Restaurantes/Café", "prioridad": "alta"},
    "NKE": {"sector": "Indumentaria Deportiva", "prioridad": "alta"},
    "T": {"sector": "Telecomunicaciones", "prioridad": "alta"},
    "VZ": {"sector": "Telecomunicaciones", "prioridad": "alta"},
    "TMUS": {"sector": "Telecomunicaciones", "prioridad": "media"},
    "VOD": {"sector": "Telecom UK", "prioridad": "media"},
    "AMX": {"sector": "Telecom México", "prioridad": "media"},
    "NOK": {"sector": "Telecom Finlandia", "prioridad": "media"},
    "ERIC": {"sector": "Telecom Suecia", "prioridad": "media"},
    "MELI": {"sector": "E-commerce LatAm", "prioridad": "muy_alta"},
    "BABA": {"sector": "E-commerce China", "prioridad": "alta"},
    "JD": {"sector": "E-commerce China", "prioridad": "alta"},
    "PDD": {"sector": "E-commerce China", "prioridad": "alta"},
    "HDB": {"sector": "Bancos India", "prioridad": "alta"},
    "IBN": {"sector": "Bancos India", "prioridad": "alta"},
    "INFY": {"sector": "IT India", "prioridad": "media"},
    "SONY": {"sector": "Electrónica/Entretenimiento", "prioridad": "alta"},
    "SAP": {"sector": "Software Alemania", "prioridad": "alta"},
    "TSM": {"sector": "Semiconductores Taiwan", "prioridad": "muy_alta"},
}


class DCFAnalyzer:
    def __init__(self, tickers_dict, projection_years=5, filters=None):
        self.tickers_dict = tickers_dict
        self.years = projection_years
        self.results = []
        self.filtered_results = []
        self.filters = filters or FILTER_CONFIG
        self.processed_count = 0
        self.checkpoint_file = "dcf_checkpoint.csv"

    # ============================================
    # MÉTODOS DE UTILIDAD
    # ============================================

    def is_bank(self, sector):
        return any(bank_sector in sector for bank_sector in BANK_SECTORS)

    def is_etf(self, ticker):
        return ticker in ETF_TICKERS

    def classify_by_market_cap(self, market_cap):
        if market_cap > 200e9:
            return "MEGA CAP (>$200B)"
        elif market_cap > 10e9:
            return "LARGE CAP ($10B-$200B)"
        elif market_cap > 2e9:
            return "MID CAP ($2B-$10B)"
        elif market_cap > 300e6:
            return "SMALL CAP ($300M-$2B)"
        else:
            return "MICRO CAP (<$300M)"

    def save_checkpoint(self):
        if (
            self.filtered_results
            and self.processed_count % self.filters["save_checkpoint_every"] == 0
        ):
            df = pd.DataFrame(self.filtered_results)
            df.to_csv(self.checkpoint_file, index=False)
            print(
                f"      💾 Checkpoint guardado: {len(self.filtered_results)} acciones"
            )

    def load_checkpoint(self):
        try:
            import os

            if os.path.exists(self.checkpoint_file):
                df = pd.read_csv(self.checkpoint_file)
                tickers = set(df["ticker"].tolist())
                print(
                    f"   📂 Checkpoint encontrado: {len(tickers)} acciones ya analizadas"
                )
                return tickers
        except Exception:
            pass
        return set()

    # ============================================
    # MÉTRICAS DE VALORACIÓN ADICIONALES
    # ============================================

    def get_valuation_metrics(self, ticker, current_price, shares_b):
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            metrics = {}

            book_value = info.get("bookValue", 0)
            if book_value and book_value > 0:
                metrics["pb_ratio"] = round(current_price / book_value, 2)

            revenue_data = self.get_revenue_growth_analysis(ticker)
            if revenue_data and revenue_data.get("has_data", False):
                latest_revenue = revenue_data.get("latest_revenue", 0)
                if latest_revenue > 0 and shares_b > 0:
                    sales_per_share = latest_revenue / shares_b
                    metrics["ps_ratio"] = round(current_price / sales_per_share, 2)

            dividend_rate = info.get("dividendRate", 0)
            if dividend_rate and dividend_rate > 0:
                metrics["dividend_yield"] = round(
                    (dividend_rate / current_price) * 100, 2
                )

            forward_pe = info.get("forwardPE", 0)
            if forward_pe and forward_pe > 0:
                metrics["forward_pe"] = round(forward_pe, 2)

            return metrics
        except Exception:
            return {}

    # ============================================
    # SCORE DE CALIDAD FINANCIERA
    # ============================================

    def get_quality_score(self, ticker):
        try:
            stock = yf.Ticker(ticker)
            financials = stock.financials
            cashflow = stock.cashflow
            balance = stock.balance_sheet

            if financials.empty or len(financials.columns) < 2:
                return 5

            score = 0

            # 1. ROA positivo
            net_income = None
            total_assets = None
            if "Net Income" in financials.index and "Total Assets" in balance.index:
                net_income = financials.loc["Net Income"].iloc[0]
                total_assets = balance.loc["Total Assets"].iloc[0]
                if net_income and total_assets and net_income > 0 and total_assets > 0:
                    score += 1

            # 2. CFO positivo
            if "Operating Cash Flow" in cashflow.index and not cashflow.empty:
                cfo = cashflow.loc["Operating Cash Flow"].iloc[0]
                if cfo and cfo > 0:
                    score += 1

            # 3. ROA mejor que año anterior
            if (
                net_income is not None
                and total_assets is not None
                and len(financials.columns) >= 2
                and "Total Assets" in balance.index
                and len(balance.columns) >= 2
            ):
                try:
                    roa_curr = net_income / total_assets if total_assets > 0 else 0
                    net_income_prev = financials.loc["Net Income"].iloc[1]
                    total_assets_prev = balance.loc["Total Assets"].iloc[1]
                    roa_prev = (
                        net_income_prev / total_assets_prev
                        if total_assets_prev > 0
                        else 0
                    )
                    if roa_curr > roa_prev:
                        score += 1
                except Exception:
                    pass

            # 4. CFO > Net Income
            if (
                "Net Income" in financials.index
                and "Operating Cash Flow" in cashflow.index
            ):
                try:
                    net_inc = financials.loc["Net Income"].iloc[0]
                    cfo_val = cashflow.loc["Operating Cash Flow"].iloc[0]
                    if cfo_val and net_inc and cfo_val > net_inc:
                        score += 1
                except Exception:
                    pass

            # 5. Deuda disminuyó
            if "Total Debt" in balance.index and len(balance.columns) >= 2:
                try:
                    debt_curr = balance.loc["Total Debt"].iloc[0]
                    debt_prev = balance.loc["Total Debt"].iloc[1]
                    if debt_curr < debt_prev:
                        score += 1
                except Exception:
                    pass

            # 6. Ratio corriente mejoró
            if (
                "Current Assets" in balance.index
                and "Current Liabilities" in balance.index
                and len(balance.columns) >= 2
            ):
                try:
                    curr_ratio = (
                        balance.loc["Current Assets"].iloc[0]
                        / balance.loc["Current Liabilities"].iloc[0]
                    )
                    curr_ratio_prev = (
                        balance.loc["Current Assets"].iloc[1]
                        / balance.loc["Current Liabilities"].iloc[1]
                    )
                    if curr_ratio > curr_ratio_prev:
                        score += 1
                except Exception:
                    pass

            # 7. Punto base por consistencia
            score += 1

            # 8. Apalancamiento
            if "Total Debt" in balance.index and "Total Equity" in balance.index:
                try:
                    debt_to_equity = (
                        balance.loc["Total Debt"].iloc[0]
                        / balance.loc["Total Equity"].iloc[0]
                    )
                    if debt_to_equity < 1:
                        score += 1
                except Exception:
                    pass

            return min(9, score)
        except Exception:
            return 5

    # ============================================
    # MÉTRICAS PARA BANCOS
    # ============================================

    def get_bank_metrics(self, ticker):
        try:
            stock = yf.Ticker(ticker)
            info = stock.info

            roe = info.get("returnOnEquity", 0)
            if roe and not np.isnan(roe):
                roe = roe * 100 if roe < 1 else roe

            tier1_ratio = info.get("tier1CapitalRatio", 0)
            if tier1_ratio and not np.isnan(tier1_ratio):
                tier1_ratio = tier1_ratio * 100 if tier1_ratio < 1 else tier1_ratio

            if roe > 15:
                bank_health = "🟢 SALUDABLE"
                health_score = 1
            elif roe > 10:
                bank_health = "🟡 ACEPTABLE"
                health_score = 2
            elif roe > 5:
                bank_health = "🟠 DÉBIL"
                health_score = 3
            else:
                bank_health = "🔴 RIESGO"
                health_score = 4

            return {
                "roe": round(roe, 1),
                "tier1": round(tier1_ratio, 1) if tier1_ratio > 0 else None,
                "bank_health": bank_health,
                "health_score": health_score,
                "is_bank": True,
            }
        except Exception:
            return {
                "is_bank": True,
                "roe": 0,
                "tier1": None,
                "bank_health": "⚪ SIN DATOS",
                "health_score": 3,
            }

    # ============================================
    # MÉTRICAS DE CRECIMIENTO
    # ============================================

    def get_revenue_growth_analysis(self, ticker):
        try:
            stock = yf.Ticker(ticker)
            financials = stock.financials

            if financials is None or financials.empty:
                return None

            if "Total Revenue" not in financials.index:
                return None

            revenues = financials.loc["Total Revenue"]

            if len(revenues) < 2:
                return None

            revenue_data = []
            years_list = []

            for i, (col, value) in enumerate(revenues.items()):
                if i >= 5:
                    break
                year = col.year if hasattr(col, "year") else str(col)[:4]
                revenue_billions = value / 1e9
                revenue_data.append(revenue_billions)
                years_list.append(year)

            revenue_data = list(reversed(revenue_data))
            years_list = list(reversed(years_list))

            yearly_growth = []
            for i in range(1, len(revenue_data)):
                if revenue_data[i - 1] > 0:
                    growth = (
                        (revenue_data[i] - revenue_data[i - 1])
                        / revenue_data[i - 1]
                        * 100
                    )
                    yearly_growth.append(round(growth, 1))
                else:
                    yearly_growth.append(0)

            cagr_3y = 0
            cagr_5y = 0

            if len(revenue_data) >= 4:
                oldest_3y = revenue_data[-4]
                newest = revenue_data[-1]
                if oldest_3y > 0:
                    cagr_3y = ((newest / oldest_3y) ** (1 / 3) - 1) * 100

            if len(revenue_data) >= 2:
                oldest = revenue_data[0]
                newest = revenue_data[-1]
                years = len(revenue_data) - 1
                if oldest > 0 and years > 0:
                    cagr_5y = ((newest / oldest) ** (1 / years) - 1) * 100

            positive_years = sum(1 for g in yearly_growth if g > 0)
            consistency = (
                round(positive_years / len(yearly_growth) * 100) if yearly_growth else 0
            )

            if cagr_5y > 20 and consistency > 70:
                growth_profile = "🚀 CRECIMIENTO ACELERADO"
            elif cagr_5y > 10 and consistency > 60:
                growth_profile = "📈 CRECIMIENTO SÓLIDO"
            elif cagr_5y > 5 and consistency > 50:
                growth_profile = "➡️ CRECIMIENTO MODERADO"
            elif cagr_5y > 0:
                growth_profile = "🐢 CRECIMIENTO LENTO"
            elif cagr_5y > -5:
                growth_profile = "📉 ESTANCAMIENTO"
            else:
                growth_profile = "⚠️ DECRECIMIENTO"

            return {
                "cagr_3y": round(cagr_3y, 1),
                "cagr_5y": round(cagr_5y, 1),
                "consistency": consistency,
                "growth_profile": growth_profile,
                "latest_revenue": round(revenue_data[-1], 1) if revenue_data else 0,
                "latest_growth": yearly_growth[-1] if yearly_growth else 0,
                "has_data": True,
            }
        except Exception:
            return {"has_data": False}

    # ============================================
    # MÉTRICAS DE DEUDA
    # ============================================

    def get_debt_metrics(self, ticker):
        try:
            stock = yf.Ticker(ticker)
            balance = stock.balance_sheet
            financials = stock.financials

            if balance.empty:
                return None

            total_debt = 0
            if "Total Debt" in balance.index:
                total_debt = balance.loc["Total Debt"].iloc[0] / 1e9

            short_term_debt = 0
            if "Short Term Debt" in balance.index:
                short_term_debt = balance.loc["Short Term Debt"].iloc[0] / 1e9

            cash = 0
            if "Cash And Cash Equivalents" in balance.index:
                cash = balance.loc["Cash And Cash Equivalents"].iloc[0] / 1e9
            if "Short Term Investments" in balance.index:
                cash += balance.loc["Short Term Investments"].iloc[0] / 1e9

            net_debt = total_debt - cash

            ebitda = 0
            if "EBITDA" in financials.index:
                ebitda = financials.loc["EBITDA"].iloc[0] / 1e9

            interest_expense = 0
            if "Interest Expense" in financials.index:
                interest_expense = abs(financials.loc["Interest Expense"].iloc[0]) / 1e9

            if ebitda > 0:
                debt_to_ebitda = total_debt / ebitda
            else:
                debt_to_ebitda = None

            interest_coverage = (
                ebitda / interest_expense if interest_expense > 0 else 999
            )

            if total_debt > 0:
                short_term_ratio = (short_term_debt / total_debt) * 100
            else:
                short_term_ratio = 0

            if debt_to_ebitda is not None:
                if (
                    debt_to_ebitda < 1.5
                    and interest_coverage > 10
                    and short_term_ratio < 20
                ):
                    risk_level = "MUY BAJO"
                    risk_color = "🟢"
                    risk_score = 1
                elif (
                    debt_to_ebitda < 3
                    and interest_coverage > 5
                    and short_term_ratio < 30
                ):
                    risk_level = "BAJO"
                    risk_color = "🟢"
                    risk_score = 2
                elif (
                    debt_to_ebitda < 5
                    and interest_coverage > 3
                    and short_term_ratio < 40
                ):
                    risk_level = "MODERADO"
                    risk_color = "🟡"
                    risk_score = 3
                elif (
                    debt_to_ebitda < 8
                    and interest_coverage > 1.5
                    and short_term_ratio < 50
                ):
                    risk_level = "ALTO"
                    risk_color = "🟠"
                    risk_score = 4
                else:
                    risk_level = "MUY ALTO"
                    risk_color = "🔴"
                    risk_score = 5
            else:
                risk_level = "N/A"
                risk_color = "⚪"
                risk_score = 3

            return {
                "total_debt_b": round(total_debt, 1),
                "net_debt_b": round(net_debt, 1),
                "debt_to_ebitda": round(debt_to_ebitda, 1) if debt_to_ebitda else None,
                "interest_coverage": round(interest_coverage, 1),
                "short_term_ratio": round(short_term_ratio, 1),
                "risk_level": risk_level,
                "risk_color": risk_color,
                "risk_score": risk_score,
            }
        except Exception:
            return None

    # ============================================
    # DATOS BÁSICOS
    # ============================================

    def get_fcf_billions(self, ticker):
        try:
            stock = yf.Ticker(ticker)
            cashflow = stock.cashflow
            if cashflow is None or cashflow.empty:
                return None
            if "Free Cash Flow" in cashflow.index:
                fcf = cashflow.loc["Free Cash Flow"].iloc[0]
                if fcf and not np.isnan(fcf):
                    return fcf / 1e9
            if (
                "Operating Cash Flow" in cashflow.index
                and "Capital Expenditure" in cashflow.index
            ):
                ocf = cashflow.loc["Operating Cash Flow"].iloc[0]
                capex = cashflow.loc["Capital Expenditure"].iloc[0]
                if ocf and capex:
                    return (ocf - abs(capex)) / 1e9
            return None
        except Exception:
            return None

    def get_shares_billions(self, ticker):
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            shares = info.get("sharesOutstanding", 0)
            if shares and not np.isnan(shares) and shares > 0:
                return shares / 1e9
        except Exception:
            pass
        return None

    def get_current_price(self, ticker):
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            price = info.get("currentPrice", info.get("regularMarketPrice", 0))
            if price and not np.isnan(price):
                return price
        except Exception:
            pass
        return 0

    def get_pe_ratio(self, ticker, current_price):
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            pe = info.get("trailingPE", 0)
            if pe and not np.isnan(pe) and pe > 0:
                return pe
        except Exception:
            pass
        return 0

    def get_market_cap(self, ticker):
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            return info.get("marketCap", 0)
        except Exception:
            return 0

    # ============================================
    # CÁLCULO DE WACC Y VALOR INTRÍNSECO
    # ============================================

    def calculate_wacc(self, sector, risk_score=2):
        wacc_map = {
            "Tecnología": 0.09,
            "Cloud/AI": 0.095,
            "Semiconductores/AI": 0.10,
            "Internet/Publicidad": 0.085,
            "Social Media": 0.09,
            "E-commerce/Cloud": 0.085,
            "Automoción/Tech": 0.095,
            "Bancos/Custodia": 0.07,
            "Banca Inversión": 0.075,
            "Bancos": 0.07,
            "Medios de Pago": 0.075,
            "Seguros": 0.08,
            "Broker": 0.075,
            "Energía (Vaca Muerta)": 0.11,
            "Energía Argentina": 0.12,
            "Petróleo & Gas": 0.085,
            "Bebidas": 0.07,
            "Consumo": 0.07,
            "Retail/Masivo": 0.075,
            "Salud": 0.075,
            "Farmacéutica": 0.075,
            "Industrial": 0.08,
            "Aeroespacial": 0.085,
            "Telecomunicaciones": 0.08,
            "ETF": 0.08,
            "General": 0.08,
        }
        base_wacc = wacc_map.get(sector, 0.08)
        risk_adjustment = (risk_score - 2) * 0.005
        return max(0.05, min(0.15, base_wacc + risk_adjustment))

    def calculate_intrinsic_value(
        self, fcf_b, growth_rate, wacc, shares_b, net_debt_b, years
    ):
        if fcf_b <= 0 or shares_b <= 0:
            return 0

        projected_fcf = []
        for i in range(1, years + 1):
            fcf = fcf_b * (1 + growth_rate) ** i
            projected_fcf.append(fcf)

        terminal_growth = 0.025
        if wacc <= terminal_growth:
            terminal_growth = wacc - 0.005
        terminal_value = (
            projected_fcf[-1] * (1 + terminal_growth) / (wacc - terminal_growth)
        )

        pv_fcf = sum(fcf / ((1 + wacc) ** t) for t, fcf in enumerate(projected_fcf, 1))
        pv_terminal = terminal_value / ((1 + wacc) ** years)
        enterprise_value = pv_fcf + pv_terminal
        equity_value = enterprise_value - net_debt_b
        price_per_share = equity_value / shares_b if shares_b > 0 else 0

        return price_per_share

    # ============================================
    # FILTROS
    # ============================================

    def should_filter(
        self, margin, cagr_5y, debt_to_ebitda, flag, is_bank, quality_score
    ):
        if margin < self.filters["min_margin"] or margin > self.filters["max_margin"]:
            return True
        if cagr_5y < self.filters["min_cagr"]:
            return True
        if not is_bank and debt_to_ebitda is not None:
            if debt_to_ebitda > self.filters["max_debt_to_ebitda"]:
                return True
        if quality_score < self.filters.get("min_quality_score", 0):
            return True
        if self.filters["show_only_buy"] and flag not in ["FUERTE_COMPRA", "COMPRA"]:
            return True
        return False

    # ============================================
    # ANÁLISIS PRINCIPAL
    # ============================================

    def analyze_ticker(self, ticker, sector):
        if self.is_etf(ticker):
            return None

        fcf_b = self.get_fcf_billions(ticker)
        if fcf_b is None or fcf_b <= 0:
            return None

        shares_b = self.get_shares_billions(ticker)
        if shares_b is None or shares_b <= 0:
            return None

        current_price = self.get_current_price(ticker)
        if current_price <= 0:
            return None

        is_bank = self.is_bank(sector)

        growth_analysis = self.get_revenue_growth_analysis(ticker)
        market_cap = self.get_market_cap(ticker)
        size_classification = self.classify_by_market_cap(market_cap)
        quality_score = self.get_quality_score(ticker)

        if growth_analysis and growth_analysis.get("has_data", False):
            if growth_analysis["cagr_5y"] > 0:
                growth_rate = max(0.01, min(0.20, growth_analysis["cagr_5y"] / 100))
            else:
                growth_rate = 0.05
            cagr_5y = growth_analysis["cagr_5y"]
            growth_profile = growth_analysis["growth_profile"]
            cagr_display = f"{cagr_5y:.1f}%"
        else:
            growth_rate = 0.05
            cagr_5y = 0
            growth_profile = "📊 DATOS PARCIALES"
            cagr_display = "N/D"

        pe_ratio = self.get_pe_ratio(ticker, current_price)
        valuation_metrics = self.get_valuation_metrics(ticker, current_price, shares_b)

        if is_bank:
            bank_metrics = self.get_bank_metrics(ticker)
            net_debt_b = 0
            risk_score = bank_metrics.get("health_score", 3)
            risk_level = bank_metrics.get("bank_health", "DESCONOCIDO")
            risk_color = (
                "🟢"
                if "SALUDABLE" in risk_level
                else "🟡"
                if "ACEPTABLE" in risk_level
                else "🟠"
                if "DÉBIL" in risk_level
                else "🔴"
            )
            total_debt_b = 0
            debt_to_ebitda = None
            roe = bank_metrics.get("roe", 0)
            tier1 = bank_metrics.get("tier1", None)
        else:
            debt_metrics = self.get_debt_metrics(ticker)
            if debt_metrics:
                net_debt_b = debt_metrics["net_debt_b"]
                risk_score = debt_metrics.get("risk_score", 2)
                risk_level = debt_metrics.get("risk_level", "DESCONOCIDO")
                risk_color = debt_metrics.get("risk_color", "⚪")
                total_debt_b = debt_metrics["total_debt_b"]
                debt_to_ebitda = debt_metrics.get("debt_to_ebitda")
            else:
                net_debt_b = 0
                risk_score = 2
                risk_level = "DESCONOCIDO"
                risk_color = "⚪"
                total_debt_b = 0
                debt_to_ebitda = None
            roe = None
            tier1 = None

        wacc = self.calculate_wacc(sector, risk_score)
        intrinsic_value = self.calculate_intrinsic_value(
            fcf_b, growth_rate, wacc, shares_b, net_debt_b, self.years
        )

        if intrinsic_value <= 0:
            return None

        margin = (intrinsic_value - current_price) / current_price * 100

        if margin > 30:
            flag = "FUERTE_COMPRA"
            emoji = "🚀"
        elif margin > 15:
            flag = "COMPRA"
            emoji = "✅"
        elif margin > 5:
            flag = "INTERESANTE"
            emoji = "📊"
        elif margin > -10:
            flag = "PRECIO_JUSTO"
            emoji = "➖"
        elif margin > -25:
            flag = "SOBREVALORADA"
            emoji = "⚠️"
        else:
            flag = "EVITAR"
            emoji = "❌"

        if self.should_filter(
            margin, cagr_5y, debt_to_ebitda, flag, is_bank, quality_score
        ):
            return None

        # Mostrar en pantalla
        if is_bank:
            tier1_str = f" | CET1: {tier1}%" if tier1 else ""
            print(
                f"   {emoji} ${intrinsic_value:.2f} | Actual: ${current_price:.2f} | Margen: {margin:+.1f}% | CAGR5: {cagr_display} | {growth_profile}"
            )
            print(
                f"      🏦 BANCO | Salud: {risk_color}{risk_level} | ROE: {roe}%{tier1_str} | Score: {quality_score}/9"
            )
        else:
            debt_str = f"{debt_to_ebitda:.1f}x" if debt_to_ebitda is not None else "N/A"
            pb_str = (
                f" | P/B: {valuation_metrics.get('pb_ratio', 'N/A')}"
                if valuation_metrics.get("pb_ratio")
                else ""
            )
            print(
                f"   {emoji} ${intrinsic_value:.2f} | Actual: ${current_price:.2f} | Margen: {margin:+.1f}% | CAGR5: {cagr_display} | {growth_profile}"
            )
            print(
                f"      📊 Deuda: Total ${total_debt_b:.1f}B | Neta ${net_debt_b:+.1f}B | D/EBITDA: {debt_str} | Riesgo: {risk_color}{risk_level}"
            )
            print(
                f"      📈 Métricas: P/E: {pe_ratio:.1f}{pb_str} | Calidad: {quality_score}/9 | Tamaño: {size_classification}"
            )

        return {
            "ticker": ticker,
            "sector": sector,
            "current_price": round(current_price, 2),
            "intrinsic_value": round(intrinsic_value, 2),
            "margin": round(margin, 1),
            "flag": flag,
            "emoji": emoji,
            "cagr_5y": cagr_5y if cagr_5y != 0 else None,
            "growth_profile": growth_profile,
            "total_debt_b": total_debt_b if not is_bank else None,
            "net_debt_b": net_debt_b if not is_bank else None,
            "debt_to_ebitda": debt_to_ebitda,
            "risk_level": risk_level,
            "is_bank": is_bank,
            "roe": roe,
            "tier1": tier1,
            "quality_score": quality_score,
            "size_classification": size_classification,
            "pe_ratio": round(pe_ratio, 1) if pe_ratio else None,
            "pb_ratio": valuation_metrics.get("pb_ratio"),
            "ps_ratio": valuation_metrics.get("ps_ratio"),
            "dividend_yield": valuation_metrics.get("dividend_yield"),
            "forward_pe": valuation_metrics.get("forward_pe"),
        }

    def analyze_ticker_parallel(self, ticker_info):
        ticker, info = ticker_info
        sector = info.get("sector", "General")
        return self.analyze_ticker(ticker, sector)

    # ============================================
    # EJECUCIÓN PRINCIPAL
    # ============================================

    def run(self):
        print("\n" + "=" * 100)
        print(" DCF ANALYSIS PROFESIONAL ".center(100, "="))
        print("=" * 100)
        print(f"\n📊 Analizando {len(self.tickers_dict)} acciones...")
        print(f"📅 Proyección a {self.years} años")
        print(f"⚡ Procesamiento paralelo con {self.filters['max_workers']} hilos")
        print("\n🔧 FILTROS ACTIVOS:")
        print(f"   • Margen mínimo: {self.filters['min_margin']}%")
        print(f"   • Margen máximo: {self.filters['max_margin']}%")
        print(f"   • CAGR mínimo: {self.filters['min_cagr']}%")
        print(
            f"   • Deuda/EBITDA máximo: {self.filters['max_debt_to_ebitda']}x (solo no bancos)"
        )
        print(f"   • Calidad mínima: {self.filters.get('min_quality_score', 0)}/9")
        if self.filters["show_only_buy"]:
            print("   • Mostrar solo: COMPRA/FUERTE_COMPRA")
        print("\n📌 NOTA: Para BANCOS se usan métricas específicas (ROE, CET1)")
        print("\n" + "-" * 100 + "\n")

        _ = self.load_checkpoint()
        self.filtered_results = []

        items = list(self.tickers_dict.items())
        total = len(items)

        with concurrent.futures.ThreadPoolExecutor(
            max_workers=self.filters["max_workers"]
        ) as executor:
            future_to_ticker = {
                executor.submit(self.analyze_ticker_parallel, item): item[0]
                for item in items
            }

            for i, future in enumerate(
                concurrent.futures.as_completed(future_to_ticker), 1
            ):
                ticker = future_to_ticker[future]
                print(f"[{i}/{total}] {ticker}", end=" ")
                try:
                    result = future.result()
                    if result:
                        self.filtered_results.append(result)
                        print("✅")
                    else:
                        print("⏭️")
                except Exception as e:
                    print(f"❌ Error: {e}")

                self.processed_count = i
                self.save_checkpoint()

        return self.filtered_results

    # ============================================
    # DASHBOARD Y REPORTES
    # ============================================

    def print_dashboard(self):
        if not self.filtered_results:
            print("\n❌ No hay resultados para mostrar")
            return

        df = pd.DataFrame(self.filtered_results)
        df = df.sort_values("margin", ascending=False)

        print("\n" + "=" * 100)
        print(" 📊 DASHBOARD DE INVERSIÓN ".center(100, "="))
        print("=" * 100)

        print(f"\n📈 RESUMEN GENERAL:")
        print(f"   • Total analizadas: {len(self.tickers_dict)}")
        print(f"   • Pasaron filtros: {len(self.filtered_results)}")
        print(
            f"   • Tasa de éxito: {len(self.filtered_results) / len(self.tickers_dict) * 100:.1f}%"
        )

        print("\n🏆 TOP 5 POR MARGEN DE SEGURIDAD:")
        top5 = df.nlargest(5, "margin")[
            ["ticker", "margin", "current_price", "intrinsic_value", "quality_score"]
        ]
        for _, row in top5.iterrows():
            print(
                f"   {row['ticker']:6} | +{row['margin']:.1f}% | ${row['current_price']:.2f} → ${row['intrinsic_value']:.2f} | Calidad: {row['quality_score']}/9"
            )

        print("\n⭐ TOP 5 POR CALIDAD FINANCIERA:")
        top_quality = df.nlargest(5, "quality_score")[
            ["ticker", "quality_score", "margin", "cagr_5y", "risk_level"]
        ]
        for _, row in top_quality.iterrows():
            cagr_val = row["cagr_5y"] if row["cagr_5y"] else 0
            print(
                f"   {row['ticker']:6} | Calidad: {row['quality_score']}/9 | Margen: {row['margin']:+.1f}% | CAGR: {cagr_val:.1f}%"
            )

        print("\n📋 DESGLOSE POR RECOMENDACIÓN:")
        counts = df["flag"].value_counts()
        for flag, count in counts.items():
            emoji = {
                "FUERTE_COMPRA": "🚀",
                "COMPRA": "✅",
                "INTERESANTE": "📊",
                "PRECIO_JUSTO": "➖",
                "SOBREVALORADA": "⚠️",
                "EVITAR": "❌",
            }.get(flag, "•")
            print(f"   {emoji} {flag}: {count}")

        print("\n💼 CARTERA SUGERIDA (Top 10 por calidad compuesta):")
        df["composite_score"] = (
            df["margin"].clip(lower=0, upper=100) * 0.5 + df["quality_score"] * 10
        )
        portfolio = df.nlargest(10, "composite_score")[
            ["ticker", "margin", "quality_score", "cagr_5y", "risk_level"]
        ]
        for _, row in portfolio.iterrows():
            cagr_val = row["cagr_5y"] if row["cagr_5y"] else 0
            print(
                f"   {row['ticker']:6} | Margen: {row['margin']:+.1f}% | Calidad: {row['quality_score']}/9 | CAGR: {cagr_val:.1f}% | Riesgo: {row['risk_level']}"
            )

        return df

    def print_summary(self):
        if not self.filtered_results:
            print("\n❌ No hay resultados después de aplicar los filtros")
            return

        df = pd.DataFrame(self.filtered_results)
        df = df.sort_values("margin", ascending=False)

        print("\n" + "=" * 100)
        print(" RESUMEN DE RESULTADOS FILTRADOS ".center(100, "="))
        print("=" * 100)

        best = df[df["flag"].isin(["FUERTE_COMPRA", "COMPRA"])].head(15)
        if not best.empty:
            print("\n🎯 TOP RECOMENDACIONES DE INVERSIÓN:")
            print("-" * 80)
            for _, row in best.iterrows():
                if row.get("is_bank", False):
                    health = f"ROE: {row.get('roe', 'N/A')}%"
                    print(
                        f"   {row['emoji']} {row['ticker']:6} | ${row['current_price']:.2f} → ${row['intrinsic_value']:.2f} | +{row['margin']:.1f}% | {health} | Calidad: {row['quality_score']}/9"
                    )
                else:
                    debt = f"D/EBITDA: {row.get('debt_to_ebitda', 'N/A')}"
                    print(
                        f"   {row['emoji']} {row['ticker']:6} | ${row['current_price']:.2f} → ${row['intrinsic_value']:.2f} | +{row['margin']:.1f}% | {debt} | Calidad: {row['quality_score']}/9"
                    )

        # Mejores por crecimiento (usando fillna para evitar errores)
        if "cagr_5y" in df.columns:
            df_cagr = df.copy()
            df_cagr["cagr_5y"] = df_cagr["cagr_5y"].fillna(0)
            best_growth = df_cagr[
                (df_cagr["flag"].isin(["FUERTE_COMPRA", "COMPRA"]))
                & (df_cagr["cagr_5y"] > 5)
            ].head(10)
            if not best_growth.empty:
                print("\n📈 TOP OPORTUNIDADES CON CRECIMIENTO REAL (>5% CAGR):")
                print("-" * 80)
                for _, row in best_growth.iterrows():
                    print(
                        f"   {row['emoji']} {row['ticker']:6} | +{row['margin']:.1f}% margen | CAGR5: {row['cagr_5y']:.1f}% | Calidad: {row['quality_score']}/9"
                    )

        return df

    # ============================================
    # EXPORTACIÓN
    # ============================================

    def export_to_excel_formatted(self, filename="dcf_analysis_formatted.xlsx"):
        if not self.filtered_results:
            print("No hay datos para exportar")
            return

        df = pd.DataFrame(self.filtered_results)
        df = df.sort_values("margin", ascending=False)

        if OPENPYXL_AVAILABLE:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Análisis DCF", index=False)

                # Aplicar formato condicional
                worksheet = writer.sheets["Análisis DCF"]

                green_fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                yellow_fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )
                red_fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )

                # Encontrar columna de margen
                margin_col = None
                for col_idx, col_name in enumerate(df.columns, 1):
                    if col_name == "margin":
                        margin_col = col_idx
                        break

                if margin_col:
                    for row in range(2, len(df) + 2):
                        margin_cell = worksheet.cell(row=row, column=margin_col)
                        margin_value = margin_cell.value

                        if margin_value and margin_value > 30:
                            for col in range(1, len(df.columns) + 1):
                                worksheet.cell(row=row, column=col).fill = green_fill
                        elif margin_value and margin_value < -25:
                            for col in range(1, len(df.columns) + 1):
                                worksheet.cell(row=row, column=col).fill = red_fill
                        elif margin_value and margin_value < -10:
                            for col in range(1, len(df.columns) + 1):
                                worksheet.cell(row=row, column=col).fill = yellow_fill

                # Ajustar anchos de columna
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            print(f"\n✅ Excel con formato exportado a '{filename}'")
        else:
            csv_filename = filename.replace(".xlsx", ".csv")
            df.to_csv(csv_filename, index=False, encoding="utf-8")
            print(f"\n✅ Resultados exportados a '{csv_filename}' (CSV)")

        return filename

    def export_csv(self, filename=None):
        if not self.filtered_results:
            print("No hay datos para exportar")
            return

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"dcf_analysis_final_{timestamp}.csv"

        df = pd.DataFrame(self.filtered_results)
        df = df.sort_values("margin", ascending=False)
        df.to_csv(filename, index=False, encoding="utf-8")
        print(f"\n✅ Resultados exportados a '{filename}'")
        return filename


# ============================================
# EJECUCIÓN PRINCIPAL
# ============================================


def main():
    print("\n" + "🔍 DCF ANALYZER PROFESIONAL - VERSIÓN CORREGIDA ".center(100, "="))

    my_filters = {
        "min_margin": -30,
        "max_margin": 100,
        "show_only_buy": False,
        "min_cagr": -10,
        "max_debt_to_ebitda": 8,
        "min_quality_score": 3,
        "max_workers": 5,
        "save_checkpoint_every": 50,
    }

    analyzer = DCFAnalyzer(ACCIONES_US_TOP, projection_years=5, filters=my_filters)

    try:
        analyzer.run()
        analyzer.print_dashboard()
        analyzer.print_summary()

        analyzer.export_to_excel_formatted("dcf_analysis_complete.xlsx")
        analyzer.export_csv()

        print("\n" + "✅ ANÁLISIS COMPLETADO ".center(100, "="))
        print("\n📌 FUNCIONALIDADES INCLUIDAS:")
        print("   • 🏦 Detección automática de BANCOS con métricas específicas")
        print("   • 📈 Score de calidad financiera (Piotroski F-Score 0-9)")
        print("   • 📏 Clasificación por tamaño de empresa")
        print("   • 💰 Métricas de valoración adicionales (P/B, P/S, Dividendos)")
        print(
            "\n📌 RECOMENDACIÓN: Combine margen positivo con calidad > 5/9 para mejor seguridad\n"
        )

    except KeyboardInterrupt:
        print("\n\n⚠️ Análisis interrumpido por el usuario")
        print(f"💾 Checkpoint guardado en {analyzer.checkpoint_file}")
        print(
            "   Puedes continuar desde donde lo dejaste ejecutando nuevamente el script"
        )


if __name__ == "__main__":
    main()
